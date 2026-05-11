import sys
import types

import openpyxl as px


def _ensure_crypto_stub():
    """
    Some environments import `functions.py` without pycryptodome installed.
    We only need mapping helpers, not encryption routines, so provide stubs.
    """
    if "Crypto" in sys.modules:
        return

    crypto_mod = types.ModuleType("Crypto")
    cipher_mod = types.ModuleType("Crypto.Cipher")
    util_mod = types.ModuleType("Crypto.Util")
    padding_mod = types.ModuleType("Crypto.Util.Padding")

    class _AESStub:
        pass

    def _noop_pad(value, *_args, **_kwargs):
        return value

    def _noop_unpad(value, *_args, **_kwargs):
        return value

    cipher_mod.AES = _AESStub
    padding_mod.pad = _noop_pad
    padding_mod.unpad = _noop_unpad

    sys.modules["Crypto"] = crypto_mod
    sys.modules["Crypto.Cipher"] = cipher_mod
    sys.modules["Crypto.Util"] = util_mod
    sys.modules["Crypto.Util.Padding"] = padding_mod


def _read_json_config_records(workbook):
    out = []

    if "JSON_Config" in workbook.sheetnames:
        ws = workbook["JSON_Config"]
        rows = list(ws.values)
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        idx = {name: i for i, name in enumerate(headers)}
        for row in rows[1:]:
            if row is None:
                continue
            rec = {}
            for name in ("Json Mapping", "Message", "Name Range Validation"):
                i = idx.get(name)
                rec[name] = row[i] if i is not None and i < len(row) else None
            out.append(rec)
        return out

    # Compatibility mode for workbook variant with lowercase json_config sheet.
    alt_sheet = None
    for name in workbook.sheetnames:
        if str(name).strip().lower() == "json_config":
            alt_sheet = name
            break

    if not alt_sheet:
        raise ValueError(
            "Neither JSON_Config nor json_config sheet found. "
            f"Available sheets: {workbook.sheetnames}"
        )

    ws = workbook[alt_sheet]
    rows = list(ws.values)
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    name_range_idx = headers.index("Name Range") if "Name Range" in headers else 0
    mapping_idx = headers.index("JSON MAPPING") if "JSON MAPPING" in headers else -1
    if mapping_idx < 0:
        raise ValueError("json_config sheet missing 'JSON MAPPING' column")

    for row in rows[1:]:
        if not row:
            continue
        json_mapping = row[mapping_idx] if mapping_idx < len(row) else None
        if not isinstance(json_mapping, str) or not json_mapping.strip():
            continue

        message = None
        for i in range(mapping_idx + 1, len(row)):
            candidate = row[i]
            if candidate is None:
                continue
            if isinstance(candidate, str) and candidate.strip() in {"", "#REF!", "None"}:
                continue
            if isinstance(candidate, str) and candidate.strip().startswith("ITR."):
                continue
            message = candidate
            break

        name_range = row[name_range_idx] if name_range_idx < len(row) else None
        out.append(
            {
                "Json Mapping": json_mapping,
                "Message": message if message is not None else "",
                "Name Range Validation": name_range,
            }
        )
    return out


def generate_legacy_exact_json(file_path: str):
    _ensure_crypto_stub()
    from .legacy_vendor import functions as functions_mod
    from .legacy_vendor import payload_generator as payload_generator_mod
    from .legacy_vendor import transform_settings as transform_settings_mod

    GeneratePayload = payload_generator_mod.GeneratePayload

    workbook = px.load_workbook(file_path, data_only=True, read_only=True)
    records = _read_json_config_records(workbook)

    itr6_obj = {}
    sep = "."
    name_range_key = {}

    for item in records:
        json_mapping = str(item.get("Json Mapping", "")) if item.get("Json Mapping") is not None else ""
        if isinstance(json_mapping, str) and json_mapping and not json_mapping.endswith("[0]"):
            value = item.get("Message", "")
            if value in transform_settings_mod.SELECT_REPLACE_VALUES:
                value = "Value not present"

            name_range_key[json_mapping] = item.get("Name Range Validation")
            if value not in ["Value not present", "", "Named Range doesn't exist"]:
                if value in transform_settings_mod.EMPTY_REPLACE_VALUES:
                    value = ""
                elif (
                    json_mapping in transform_settings_mod.TRANSFORM_FIELDS
                    and isinstance(value, str)
                    and "-" in value
                ):
                    value = value.split("-")[0].strip()
                if " 00:00:00" in str(value):
                    value = value.strftime("%Y-%m-%d")

                if isinstance(value, str) and value.strip() == "":
                    value = value.strip()

                functions_mod.set_mapping_node_inplace(itr6_obj, json_mapping, sep, value)
        elif json_mapping:
            try:
                value_range = item.get("Name Range Validation", "")
                column_range = item.get("Message", "")
                if column_range:
                    value_range_data = workbook.defined_names[value_range]
                    column_range_data = workbook.defined_names[column_range]
                    json_mapping = json_mapping.replace("[0]", "")
                    normalized = json_mapping[:-1] if json_mapping.endswith(".") else json_mapping
                    name_range_key[normalized] = value_range
                    list_data = functions_mod.fetch_list(
                        value_range_data, column_range_data, workbook, value_range, column_range
                    )
                    if list_data:
                        functions_mod.set_mapping_node_inplace(itr6_obj, json_mapping, sep, list_data)
            except Exception:
                pass

    mat, matc = GeneratePayload.MAT_modification(workbook, itr6_obj)
    if "ITR" in itr6_obj and "ITR6" in itr6_obj["ITR"] and "ScheduleHP" in itr6_obj["ITR"]["ITR6"]:
        del itr6_obj["ITR"]["ITR6"]["ScheduleHP"]

    GeneratePayload.generate_scheduleHP(itr6_obj, workbook)
    try:
        itr6 = itr6_obj.get("ITR", {}).get("ITR6", {})
        schedule_cg = itr6.get("ScheduleCG", {})
        long_term_cg = schedule_cg.get("LongTermCapGain", {})
        nri_node = long_term_cg.get("NRIOnSec112and115")
        if nri_node:
            output_json = GeneratePayload.transform_nri_json(nri_node)
            itr6_obj["ITR"]["ITR6"]["ScheduleCG"]["LongTermCapGain"]["NRIOnSec112and115"] = output_json
    except Exception:
        pass

    GeneratePayload.modify_payload(itr6_obj)
    parent_node = GeneratePayload.parent_gen_payload(itr6_obj["ITR"]["ITR6"])
    final_payload = {"ITR": {"ITR6": parent_node}}

    from .legacy_business_errors import run_legacy_business_error_pipeline

    business_report: dict = {}
    try:
        business_report = run_legacy_business_error_pipeline(
            final_payload, workbook, mat, matc
        )
    except Exception as exc:
        business_report = {
            "business_errors": [],
            "business_error_count": 0,
            "business_validation_status_failed": False,
            "business_errors_skipped": True,
            "business_errors_skip_reason": str(exc),
        }

    try:
        workbook.close()
    except Exception:
        pass

    out = {
        "source_payload_from_excel": itr6_obj,
        "payload": final_payload,
        "mat_removed": bool(mat or matc),
        "mat": mat,
        "matc": matc,
    }
    out.update(business_report)
    return out
