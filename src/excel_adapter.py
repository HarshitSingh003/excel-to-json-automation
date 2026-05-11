from openpyxl import load_workbook


INVALID_VALUES = {"Value not present", "Named Range doesn't exist", ""}


def _safe_value(value):
    if value in INVALID_VALUES:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in INVALID_VALUES:
            return None
        return stripped
    return value


def _looks_like_route_token(value):
    return isinstance(value, str) and value.startswith("ITR.")


def _is_unusable(value):
    if value is None:
        return True
    if isinstance(value, str):
        v = value.strip()
        if v in {"", "#REF!", "None", "Named Range doesn't exist", "Value not present"}:
            return True
    return False


def _build_json_config_map(workbook):
    """
    Build mapping from JSON_Config sheet:
    Json Mapping -> Message
    """
    try:
        ws = workbook["JSON_Config"]
    except Exception:
        return {}

    rows = list(ws.values)
    if not rows:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    mapping_idx = headers.index("Json Mapping") if "Json Mapping" in headers else -1
    message_idx = headers.index("Message") if "Message" in headers else -1
    name_range_idx = (
        headers.index("Name Range Validation") if "Name Range Validation" in headers else -1
    )
    if mapping_idx < 0 or message_idx < 0:
        return {}

    result = {}
    for row in rows[1:]:
        if row is None:
            continue
        json_mapping = row[mapping_idx] if mapping_idx < len(row) else None
        message = row[message_idx] if message_idx < len(row) else None
        name_range = row[name_range_idx] if name_range_idx >= 0 and name_range_idx < len(row) else None
        if isinstance(json_mapping, str) and json_mapping.strip():
            value = _safe_value(message)
            if value is None or _looks_like_route_token(value):
                if isinstance(name_range, str) and name_range.strip():
                    value = _read_named_range_value(workbook, name_range.strip())
            result[json_mapping.strip()] = _safe_value(value)
    return result


def _build_json_config_alt_map(workbook):
    """
    Support workbook format with lowercase `json_config` sheet:
    columns like [Name Range, JSON MAPPING, ..., value_col].
    We map JSON MAPPING -> first non-empty value from remaining columns.
    """
    target_sheet = None
    for name in workbook.sheetnames:
        if str(name).strip().lower() == "json_config":
            target_sheet = name
            break
    if not target_sheet:
        return {}

    ws = workbook[target_sheet]
    rows = list(ws.values)
    if not rows:
        return {}

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    mapping_idx = header.index("JSON MAPPING") if "JSON MAPPING" in header else -1
    if mapping_idx < 0:
        return {}

    result = {}
    quality = {}
    for row in rows[1:]:
        if not row or mapping_idx >= len(row):
            continue
        mapping = row[mapping_idx]
        if not isinstance(mapping, str) or not mapping.strip():
            continue

        value = None
        # Prefer usable non-route value from trailing columns.
        for idx in range(mapping_idx + 1, len(row)):
            candidate = _safe_value(row[idx])
            if _is_unusable(candidate):
                continue
            if _looks_like_route_token(candidate):
                continue
            value = candidate
            break

        # Fallback: allow route token only if no better candidate exists.
        if value is None:
            for idx in range(mapping_idx + 1, len(row)):
                candidate = _safe_value(row[idx])
                if _is_unusable(candidate):
                    continue
                value = candidate
                break

        if value is not None:
            key = mapping.strip()
            score = 2 if not _looks_like_route_token(value) else 1
            if key not in result or score > quality.get(key, 0):
                result[key] = value
                quality[key] = score
    return result


def _read_named_range_value(workbook, named_range):
    try:
        value_range_data = workbook.defined_names[named_range]
        destinations = list(value_range_data.destinations)
        sheet_name, cell_ref = destinations[0]
        sheet = workbook[sheet_name] if isinstance(sheet_name, str) else sheet_name
        return _safe_value(sheet[cell_ref].value)
    except Exception:
        return None


def _read_cell_value(workbook, sheet_name, cell_ref):
    try:
        return _safe_value(workbook[sheet_name][cell_ref].value)
    except Exception:
        return None


def _read_reference_value(workbook, label):
    try:
        ws = workbook["Reference and name range"]
    except Exception:
        return None

    for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        if len(row) < 3:
            continue
        if isinstance(row[1], str) and row[1].strip().lower() == label.strip().lower():
            return _safe_value(row[2])
    return None


def parse_excel_to_source_payload(file_path):
    # Strict dynamic extraction: no static key mapping fallback.
    from .legacy_exact_from_file import generate_legacy_exact_json

    try:
        legacy_exact = generate_legacy_exact_json(file_path)
        source_payload = legacy_exact.get("source_payload_from_excel")
        if isinstance(source_payload, dict) and source_payload:
            return source_payload
    except Exception:
        # Continue to dynamic JSON_Config-driven fallback.
        pass

    # Dynamic fallback: build source payload from JSON_Config mappings without
    # any hardcoded field list. This keeps adapter flexible for new keys.
    wb = load_workbook(file_path, data_only=True)
    try:
        json_map = _build_json_config_map(wb)
        if not json_map:
            json_map = _build_json_config_alt_map(wb)

        if not json_map:
            raise ValueError("No JSON_Config/json_config mappings found in workbook.")

        out = {}

        def set_path(root, path, value):
            if not isinstance(path, str) or not path.strip():
                return
            clean_path = path.replace("[0]", "").strip(".")
            if not clean_path:
                return
            parts = [p for p in clean_path.split(".") if p]
            if not parts:
                return
            cur = root
            for part in parts[:-1]:
                if part not in cur or not isinstance(cur[part], dict):
                    cur[part] = {}
                cur = cur[part]
            cur[parts[-1]] = value

        for mapping, value in json_map.items():
            safe = _safe_value(value)
            if safe is None:
                continue
            # Skip route-token placeholders like "ITR....".
            if _looks_like_route_token(safe):
                continue
            set_path(out, mapping, safe)

        if not out:
            raise ValueError("Dynamic extraction produced empty payload from JSON mappings.")
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass
